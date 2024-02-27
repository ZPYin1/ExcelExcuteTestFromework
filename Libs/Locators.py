# -*- coding: utf-8 -*-
# time: 2024/1/3 9:48
# file: Locators.py
# author: ZPYin
import logging
import os.path

from openpyxl.reader.excel import load_workbook

from Libs.Configs import LOCATOR_FILES_PATH

_logger = logging.getLogger(__name__)


class Locator:
    """
    定位器
    存放网页定位的最小单元
    """

    def __init__(self, row):
        self.name = row[0]
        self.by = row[1]
        self.value = row[2]
        self.comment = row[3]

    def __call__(self, arg, *args):
        row = [
            self.name,
            self.by,
            self.value % (arg, *args),
            self.comment
        ]
        return Locator(row)


def get_sheet_names_from_excel(excel_path):
    try:
        wb = load_workbook(excel_path)
        return wb.sheetnames
    finally:
        wb.close()


def get_locators_from_excel(excel_path, sheet_name):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    locators = dict()
    for row in ws.values:
        if row[0] is None:
            continue
        locator = Locator(row)
        locators[locator.name] = Locator(row)
    return locators


class _LocatorWorksheet:
    """
    定位器工作表
    存放网页定位整个表的数据 Dict(locator.name:locator)
    """

    def __init__(self, excel_path, sheet_name):
        self.__locators = get_locators_from_excel(excel_path, sheet_name)
        for name, value in self.__locators.items():
            self.__setattr__(name, value)

    def __call__(self, item):
        return self.__locators[item]


class _LocatorWorkbook:
    """
    定位器工作簿
    存放网页定位整个工作簿的数据 Dict(sheet_name:LocatorWorksheet)
    """

    def __init__(self, excel_path):
        for sheet_name in get_sheet_names_from_excel(excel_path):
            self.__setattr__(sheet_name, _LocatorWorksheet(excel_path, sheet_name))


BaiDu = _LocatorWorkbook(excel_path=os.path.join(LOCATOR_FILES_PATH, "baidu.xlsx"))

if __name__ == '__main__':
    print(BaiDu.HomePage.Logo.value)
    print(BaiDu.HomePage.case_number("aaa").value)
    print(BaiDu.HomePage.case_number.name)
    print(BaiDu.HomePage.case_number.by)
    print(BaiDu.HomePage.case_number.value)
    print(os.path.join(LOCATOR_FILES_PATH, "baidu.xlsx"))
