#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2024/4/25 15:33
# @Author  : ZP Y
# @File    : Utilities.py
# @Software: PyCharm
import io
import logging
import os.path
import re
import time
import requests

# import cx_Oracle
# import xlsxwriter as xw

import pytesseract
from PIL import Image
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from Libs.Configs import Environment
# from Libs.Configs import DATA_FILES_PATH
from Libs.DataBase import SQLset

# from imported import ENV

_logger = logging.getLogger(__name__)
PYTHON_REPLACE_REGEX = re.compile(r"\w")
ALPHA_REGEX = re.compile(r"^\d+_*")

join = os.path.join


def make_python_name(string):
    """make python attribute name out of given string"""
    string = re.sub(PYTHON_REPLACE_REGEX, "", string.replace(" ", "_"))
    return re.sub(ALPHA_REGEX, "", string).lower()


def data_from_excel(excel_path, sheet_name):
    """
    功能：从 Excel 中读取测试数据
    :param excel_path: Excel 地址
    :param sheet_name: 工作表名
    :return: title, list(row,value)
    """
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    rows = [row for row in ws.values if row]
    title = rows.pop(0)[1:]
    wb.close()
    data = list()
    for row in rows:
        status = row[0]
        if status in ["Skip", None]:
            continue
        if len(row) == 2:
            data.append(row[1])
        else:
            data.append(row[1:])
    return ",".join(title), data


def data_from_excel_dict(excel_path, sheet_name):
    """
    功能：从 Excel 中读取测试数据
    :param excel_path: Excel 地址
    :param sheet_name: 工作表名
    :return: title, list(row,value)
    """
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    datas = list(ws.values)
    # print(datas)

    '''
    全部读取的数据转化为项目中需要的格式数据
    '''
    # 第一步，取出标题
    title = datas[0]

    # 第二步，再一次获取后续每一行数据，跟title压缩成字典
    case_list = []
    for case in datas[1:]:
        case_dict = dict(zip(title, case))
        if case_dict['status'] in ["Skip", None]:
            continue
        case_list.append(case_dict)
    # print(case_list)
    return case_list


def data_from_text(file_path):
    with open(file_path) as f:
        return [line.strip() for line in f.readlines()]


def image_to_string(binary_data=None, filename=None, lang=None, config='', nice=0, timeout=0):
    if binary_data is None and filename is None:
        raise ValueError("binary_data or filename must exist one")
    if binary_data:
        image = Image.open(io.BytesIO(binary_data))
    elif filename:
        image = Image.open(filename)
    string = pytesseract.image_to_string(image=image, lang=lang, config=config, nice=nice, timeout=timeout)
    return string.strip()


def get_line_value(line, sep=" "):
    print(line)
    line = line[line.index(sep) + 1:]
    return line.strip()


def generate_trace_spreadsheet(test_cases, save_path):
    wb = Workbook()
    test_case_sheet = wb.create_sheet("TestCases", 0)
    test_case_sheet.append(["Name", "Author", "Traced Requirement", "Description", "Last update"])
    req_to_tc = dict()
    for tc in test_cases:
        test_case_sheet.append(tc.get_row_data())
        for req in tc.get_requirements():
            if req not in req_to_tc:
                req_to_tc[req] = list()
            req_to_tc[req].append(tc.get_id())
    req_trace_sheet = wb.create_sheet("RequirementsTrace", 1)
    req_trace_sheet.append(["Requirement ID", "Test Cases"])
    for req, tc in req_to_tc.items():
        req_trace_sheet.append([req, "\n".join(tc)])
    wb.save(save_path)


# def get_timestamp(time_fmt='%Y_%m_%d-%H_%M_%S', t=None):
def get_timestamp(time_fmt='%Y_%m_%d-%H', t=None):
    """
    获取带日期的时间格式
    :param time_fmt:
    :param t:
    :return:
    """
    t = t if t else time.time()
    return time.strftime(time_fmt, time.localtime(t))


def get_time(time_fmt='%H:%M:%S', t=None):
    """
    获取当前时间
    """
    t = t if t else time.time()
    return time.strftime(time_fmt, time.localtime(t))


class GetTestCases:
    """
    用作获取需要测试的测试用例名单
    from_excel: 从excel中获取
    from_xml:
    from_url:
    from_xxx
    返回：List(test_case)
    """

    @staticmethod
    def from_excel(excel_path):
        test_case = list()
        wb = load_workbook(excel_path, read_only=True)
        ws = wb["TestCases"]
        rows = [row for row in ws.values if row]
        title = rows.pop(0)
        run_index = title.index("Run")
        path_index = title.index("Automation Path")
        for row in rows:
            if row[run_index] == "Yes":
                test_case.append(row[path_index])
        return test_case


def xw_toExcel(fileName):
    if os.path.exists(fileName):
        workbook = load_workbook(fileName)
    else:
        workbook = Workbook(fileName)  # 创建工作簿
    ws = workbook.active  # 激活 worksheet
    workbook.save(fileName)


def data_from_SQL(env, db, sql, keyname=None, keynum=None, *args):
    result = []
    conn = SQLset(env, db)
    res_list = conn.fetchall(sql)
    del conn
    if res_list == None:
        raise Exception(f"Database no data with {sql}")
    if type(keynum) == list:
        if len(keynum) != len(keyname):
            raise Exception("key list not equal request data length")
    if keynum:
        for rs in res_list:
            if type(keynum) == list:
                temp = []
                for j in keynum:
                    temp.append(rs[j])
                if keyname:
                    temp = dict(zip(keyname, temp))
                result.append(temp)
            else:
                result.append(rs[keynum])
    else:
        for rs in res_list:
            result.append(list(rs))
    # for i in res_list:
    #     res.append(i[keynum])
    # for i in keynum:
    #     print(i)

    return result


class RequestHandler:
    def __init__(self):
        """session管理器"""
        self.session = requests.session()

    def visit(self, method, url, params=None, data=None, json=None, headers=None, **kwargs):
        return self.session.request(method, url, params=params, data=data, json=json, headers=headers, **kwargs)

    def close_session(self):
        """关闭session"""
        self.session.close()


import platform


def get_operating_system():
    system = platform.system()

    if system == 'Windows':
        return 'Windows'
    elif system == 'Darwin':
        return 'MacOS'
    elif system == 'Linux':
        return 'Linux'
    else:
        return 'Unknown'


def generate_random_chinese_characters(num_chars):
    # reload(sys)
    # sys.setdefaultencoding("utf8")
    arr = ''
    for i in range(num_chars):
        # if i %20 ==0 and i >0:
        #     arr = arr + '\n'
        #     continue
        import random
        arr = arr + (chr(random.randint(0x4e00, 0x9fa5)))
    # print(len(arr))
    return arr


if __name__ == '__main__':
    ENV = Environment("test", )
    print(ENV.data_path)
    # data = data_from_excel_dict(ENV.data_path, "baidu")
    # print(data)
    # print(temp)

    # userlist = data_from_excel_dict(ENV.data_path, "user")
    # for user in userlist:
    #     # print(user['status'], user['username'], user['pwd'], user['email'], user['cmpname'])
    #     print(user)

    # res = tranf_str('@', 'abc', 'def')
    # print(res)

    # xw_toExcel(os.path.join(DATA_FILES_PATH, "test.xlsx"))

    # res = data_from_SQL(db='TRAS', sql="select * from PRODUCT_INFO", keynum=2)
    # res = data_from_SQL(db='TRAS', sql="select * from PRODUCT_INFO", )
    # cx_Oracle.init_oracle_client("F:\\Oracle\\instantclient-basic-windows.x64-11.2.0.4.0\\instantclient_11_2")
    # res = data_from_SQL(db='TRAS', sql="select * from PRODUCT_INFO WHERE ROWNUM <= 50", keyname=['zh', 'en'], keynum=[2, 3])
    # res = data_from_SQL(db='TRAS', sql="select * from PRODUCT_INFO WHERE ROWNUM <= 50", keynum=2)
    # for i in res:
    # print(type(i))
    # print(i)
    # break
    # print(res)

    # print(get_operating_system())

    # print(get_line_value("Author:ZPY", ":"))
    # sline = "Author:ZPY"
    # ssep = ':'
    # print(sline.index(ssep))
    # sline = sline[sline.index(ssep) + 1:]
    # print(sline)
    # print(sline.index(ssep))

    # str1 = "Hello, world!"
    #
    # index = str1.index(",")
    # str1 = str1[str1.index(",")+1:]
    # str1 = str1.strip()
    #
    # print(index)  # 输出：7
    # print(str1)

    # url = 'http://192.168.1.150:8132/Join.aspx?action=getVC&userid=test001%40123.com'
    # req = RequestHandler()
    # res = req.visit("get", url)
    # print(res.text)
    # print(type(res))
    # print(type(res.__str__()))
    #
    # print(url.replace('%40', '@'))
    # pass
